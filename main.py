import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side
import os
import sys
import json
import re


def load_config():
    if not os.path.exists('config.json'):
        print("ERROR: No se encuentra config.json")
        sys.exit(1)
    with open('config.json', 'r') as f:
        return json.load(f)


def get_folio_number(val):
    """Extrae el número de folio de un string."""
    try:
        match = re.search(r'\d+', str(val))
        return int(match.group()) if match else None
    except:
        return None


def validate_config(config):
    if 'almacenes' not in config:
        print("ERROR: config.json debe contener 'almacenes'")
        sys.exit(1)

    for block in config['almacenes']:

        # Campos obligatorios
        if 'nombre' not in block or 'ref_banco' not in block or 'rango_folios' not in block:
            print(f"ERROR: Bloque inválido: {block}")
            sys.exit(1)

        # ref_banco debe ser lista
        if not isinstance(block['ref_banco'], list):
            print(f"ERROR: {block['nombre']} -> 'ref_banco' debe ser lista")
            sys.exit(1)

        # rango_folios debe ser lista de 2 elementos
        rango = block['rango_folios']
        if not isinstance(rango, list) or len(rango) != 2:
            print(f"ERROR: {block['nombre']} -> 'rango_folios' inválido")
            sys.exit(1)


def process_conciliacion():
    config = load_config()
    validate_config(config)

    input_banco = 'input/banco.xlsx'
    input_contabilidad = 'input/contabilidad.xlsx'
    output_file = 'output/conciliacion.xlsx'

    # Validación de archivos
    if not os.path.exists(input_banco) or not os.path.exists(input_contabilidad):
        print("ERROR: Faltan archivos de entrada en /input")
        input("Presione Enter para salir...")
        sys.exit(1)

    if not os.path.exists('output'):
        os.makedirs('output')

    print("Cargando datos...")

    try:
        df_banco = pd.read_excel(input_banco, header=None)
        df_conta = pd.read_excel(input_contabilidad, header=None)
    except Exception as e:
        print(f"ERROR al leer archivos: {e}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliación"

    current_row = 1

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    all_blocks = config['almacenes']

    for block in all_blocks:
        nombre = block['nombre']
        print(f"Procesando {nombre}...")

        # ---------------- BANCO ----------------
        mask_banco = df_banco[1].astype(str).apply(
            lambda x: any(ref in x for ref in block['ref_banco'])
        )
        data_banco = df_banco[mask_banco].copy()

        # Ordenar por fecha descendente
        try:
            data_banco[0] = pd.to_datetime(data_banco[0])
            data_banco = data_banco.sort_values(by=0, ascending=False)
        except:
            pass

        total_banco = data_banco[5].sum()

        # ---------------- CONTABILIDAD ----------------
        def in_range(val):
            f = get_folio_number(val)
            return f is not None and block['rango_folios'][0] <= f <= block['rango_folios'][1]

        mask_conta = df_conta[2].apply(in_range)
        data_conta = df_conta[mask_conta].copy()

        total_conta = data_conta[7].sum()

        # ---------------- ESCRITURA ----------------
        ws.cell(row=current_row, column=1, value=nombre).font = Font(bold=True, size=14)
        current_row += 1

        # Headers banco
        headers_banco = ["Fecha", "Concepto", "Ref", "Ref Amp", "Cargo", "Abono", "Saldo"]
        for i, h in enumerate(headers_banco, 1):
            cell = ws.cell(row=current_row, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Headers contabilidad
        ws.cell(row=current_row, column=9, value="Referencia Conta").font = header_font
        ws.cell(row=current_row, column=9).fill = header_fill
        ws.cell(row=current_row, column=10, value="Importe Conta").font = header_font
        ws.cell(row=current_row, column=10).fill = header_fill

        current_row += 1
        start_data_row = current_row

        # Datos banco
        for r_idx, row_data in enumerate(data_banco.values):
            for c_idx, value in enumerate(row_data[:7]):
                ws.cell(row=start_data_row + r_idx, column=c_idx + 1, value=value)

        # Datos contabilidad
        for r_idx, row_data in enumerate(data_conta.values):
            ws.cell(row=start_data_row + r_idx, column=9, value=row_data[2])
            ws.cell(row=start_data_row + r_idx, column=10, value=row_data[7])

        max_rows = max(len(data_banco), len(data_conta))
        current_row = start_data_row + max_rows + 1

        # Totales
        ws.cell(row=current_row, column=5, value="TOTAL BANCO:").font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=total_banco).font = Font(bold=True)

        ws.cell(row=current_row + 1, column=9, value="TOTAL CONTA:").font = Font(bold=True)
        ws.cell(row=current_row + 1, column=10, value=total_conta).font = Font(bold=True)

        diff = total_banco - total_conta

        current_row += 3
        ws.cell(row=current_row, column=1, value="DIFERENCIA:").font = Font(bold=True)
        diff_cell = ws.cell(row=current_row, column=2, value=diff)
        diff_cell.font = Font(bold=True, color="FF0000" if abs(diff) > 1 else "00B050")

        current_row += 20

    # Ajuste de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    print(f"Guardando reporte en {output_file}...")
    wb.save(output_file)
    print("Proceso completado correctamente.")


if __name__ == "__main__":
    try:
        process_conciliacion()
    except Exception as e:
        print(f"FATAL ERROR: {e}")
    input("Presione Enter para cerrar...")